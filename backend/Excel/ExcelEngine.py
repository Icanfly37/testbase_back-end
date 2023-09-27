from openpyxl import *
from Excel.rows_and_cols import *

class Excel():
    def __init__(self,path): # path are bytes not directory names
        #self.path = BytesIO(path)
        self.path = path
        self.error = 0
        # self.rows = {}
        # self.cols = {}
        self.all_object = []
    def openfile(self):
        try:
            self.excel = load_workbook(self.path)
            return self.error
        except FileNotFoundError:
            return self.error
    def allsheets(self):
        return self.excel.sheetnames
    def getrows(self):
        sheet = self.excel[self.excel.sheetnames[0]]
        num_rows = sheet.max_row # จำนวนสมาชิก row สูงสุดใน sheet
        num_cols = sheet.max_column # จำนวนสมาชิก column สูงสุดใน sheet 
        for i in range(1,num_rows+1):
            if i == 1:
                header = each_row(sheet,i,num_rows,num_cols)
            else:
                slave = each_row(sheet,i,num_rows,num_cols)
                self.all_object.append(sub_object_send(str(i-1),header,slave))
        return self.all_object
    # def getcolumns(self):
    #     sheet = self.excel[self.excel.sheetnames[0]]
    #     num_rows = sheet.max_row # จำนวนสมาชิก row สูงสุดใน sheet
    #     num_cols = sheet.max_column # จำนวนสมาชิก column สูงสุดใน sheet
    #     for i in range(1, num_cols+1):
    #         self.cols.update({"col"+str(i):each_column(sheet,i,num_cols,num_rows)}) #each_column send list datatype and collect in self.cols dict
    #     return self.cols
    def closefile(self):
        self.excel.close()