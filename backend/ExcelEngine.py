from openpyxl import *
from io import *
from rows_and_cols import *

#excel = load_workbook(BytesIO(file))

class Excel():
    def __init__(self,path): # path are bytes not directory names
        self.path = BytesIO(path)
        self.error = 0
        self.rows = {}
        self.cols = {}
    def openfile(self):
        try:
            self.excel = load_workbook(self.path)
            return self.error
        except FileNotFoundError:
            return self.error
    def allsheets(self):
        return self.excel.sheetnames
    def getrows(self):
        sheet = self.excel[self.excel.sheetnames[0]]
        num_rows = sheet.max_row # จำนวนสมาชิก row สูงสุดใน sheet
        num_cols = sheet.max_column # จำนวนสมาชิก column สูงสุดใน sheet
        for i in range(1,num_rows+1):
            self.rows.update({"row "+str(i):each_row(sheet,i,num_rows,num_cols)}) #each_row send list datatype and collect in self.rows dict
        return self.rows
    def getcolumns(self):
        sheet = self.excel[self.excel.sheetnames[0]]
        num_rows = sheet.max_row # จำนวนสมาชิก row สูงสุดใน sheet
        num_cols = sheet.max_column # จำนวนสมาชิก column สูงสุดใน sheet
        for i in range(1, num_cols+1):
            self.cols.update({"col"+str(i):each_column(sheet,i,num_cols,num_rows)}) #each_column send list datatype and collect in self.cols dict
        return self.cols
    def closefile(self):
        self.excel.close()
        
#excel = load_workbook(BytesIO(file))