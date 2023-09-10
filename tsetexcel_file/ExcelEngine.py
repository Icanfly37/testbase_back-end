from openpyxl import *

class Excel():
    def __init__(self,path):
        self.path = path
        self.error = 0
        self.rows = {}
        self.cols = {}
        try:
            self.excel = load_workbook(self.path)
        except FileNotFoundError:
            return self.error
    def allsheets(self):
        return self.excel.sheetnames
    def getrows(self):
        sheet = self.excel[self.excel.sheetnames[0]]
        i = 0
        for row in sheet.iter_rows(values_only=True):
            # print(list(row))
            self.rows.update({"row"+str(i):list(row)})
            i+=1
        return self.rows
    def getcolumns(self):
        sheet = self.excel[self.excel.sheetnames[0]]
        i = 0
        for col in sheet.iter_cols(values_only=True):
            # print(list(col))
            self.cols.update({"col"+str(i):list(col)})
            i+=1
        return self.cols
    def closefile(self):
        self.excel.close()
            
ex = Excel("D:/excel_test/test.xlsx")
print(ex.allsheets())
print(ex.getrows())
print(ex.getcolumns())
ex.closefile()